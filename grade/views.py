from django.shortcuts import render, redirect, get_object_or_404
from django.forms import modelformset_factory
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import io
from .models import Grade
from classroom.models import Classroom
from courses.models import Subject
from students.models import Student
from .form import GradeForm , CustomPrintForm
import openpyxl
from openpyxl.styles import Font, Alignment 

def grades_dashboard(request):
    classrooms = Classroom.objects.all()
    return render(request, 'grade/dashboard.html', {'classrooms': classrooms})

def view_grades(request, classroom_id, subject_id):
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    
    # الحصول على أسماء المعلمين للمادة
    teacher_names = ", ".join([teacher.full_name for teacher in subject.teachers.all()])
    subject_display_name = f"{subject.name} ({teacher_names})" if teacher_names else subject.name
    
    grades = Grade.objects.filter(
        classroom=classroom,
        subject=subject
    ).select_related('student')
    
    # حساب المجموع لكل طالب
    students_data = []
    for student in classroom.students.all().order_by('full_name'):
        student_grades = grades.filter(student=student)
        activity_grade = student_grades.filter(exam_type='activity').first()
        monthly_grade = student_grades.filter(exam_type='monthly').first()
        midterm_grade = student_grades.filter(exam_type='midterm').first()
        final_grade = student_grades.filter(exam_type='final').first()
        
        # حساب المجموع (يمكن تعديل طريقة الحساب حسب النظام)
        total = 0
        if activity_grade and activity_grade.grade: 
            total += float(activity_grade.grade)
        if monthly_grade and monthly_grade.grade: 
            total += float(monthly_grade.grade)
        if midterm_grade and midterm_grade.grade: 
            total += float(midterm_grade.grade)
        if final_grade and final_grade.grade: 
            total += float(final_grade.grade)
            
        students_data.append({
            'student': student,
            'activity': activity_grade,
            'monthly': monthly_grade,
            'midterm': midterm_grade,
            'final': final_grade,
            'total': total
        })
    
    return render(request, 'grade/view_grades.html', {
        'classroom': classroom,
        'subject': subject,
        'subject_display_name': subject_display_name,
        'students_data': students_data
    })

def edit_grades(request, classroom_id, subject_id):
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    
    # الحصول على أسماء المعلمين للمادة
    teacher_names = ", ".join([teacher.full_name for teacher in subject.teachers.all()])
    subject_display_name = f"{subject.name} ({teacher_names})" if teacher_names else subject.name
    
    GradeFormSet = modelformset_factory(Grade, form=GradeForm, extra=0)
    students = classroom.students.all().order_by('full_name')
    
    if request.method == 'POST':
        formset = GradeFormSet(request.POST, request.FILES)
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.classroom = classroom
                instance.subject = subject
                # إذا كانت العلامة فارغة، ضعها كقيمة فارغة بدلاً من 0
                if instance.grade == 0:
                    instance.grade = None
                instance.save()
            return redirect('grade:view_grades', classroom.id, subject.id)
    else:
        # جلب العلامات الموجودة للمادة الحالية فقط
        grades = Grade.objects.filter(
            classroom=classroom,
            subject=subject
        )
        
        # إنشاء علامات للطلاب المفقودين فقط
        existing_student_ids = grades.values_list('student_id', flat=True)
        missing_students = students.exclude(id__in=existing_student_ids)
        
        new_grades = []
        for student in missing_students:
            for exam_type in ['activity', 'monthly', 'midterm', 'final']:
                new_grades.append(Grade(
                    student=student,
                    classroom=classroom,
                    subject=subject,
                    exam_type=exam_type,
                    grade=None,  # بدلاً من 0، نستخدم None للقيم الفارغة
                    notes=''
                ))
        
        if new_grades:
            Grade.objects.bulk_create(new_grades)
            grades = Grade.objects.filter(classroom=classroom, subject=subject)
        
        formset = GradeFormSet(queryset=grades)
    
    return render(request, 'grade/edit_grades.html', {
        'classroom': classroom,
        'subject': subject,
        'subject_display_name': subject_display_name,
        'formset': formset,
        'students': students
    })

def print_grades(request, classroom_id, subject_id):
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    
    # الحصول على أسماء المعلمين للمادة
    teacher_names = ", ".join([teacher.full_name for teacher in subject.teachers.all()])
    subject_display_name = f"{subject.name} ({teacher_names})" if teacher_names else subject.name
    
    grades = Grade.objects.filter(
        classroom=classroom,
        subject=subject
    ).select_related('student')
    
    # أنواع الامتحانات
    exam_types = Grade.ExamType.choices
    
    # حساب المجموع لكل طالب
    students_data = []
    for student in classroom.students.all().order_by('full_name'):
        student_grades = grades.filter(student=student)
        activity_grade = student_grades.filter(exam_type='activity').first()
        monthly_grade = student_grades.filter(exam_type='monthly').first()
        midterm_grade = student_grades.filter(exam_type='midterm').first()
        final_grade = student_grades.filter(exam_type='final').first()
        
        # حساب المجموع
        total = 0
        if activity_grade and activity_grade.grade: 
            total += float(activity_grade.grade)
        if monthly_grade and monthly_grade.grade: 
            total += float(monthly_grade.grade)
        if midterm_grade and midterm_grade.grade: 
            total += float(midterm_grade.grade)
        if final_grade and final_grade.grade: 
            total += float(final_grade.grade)
            
        students_data.append({
            'student': student,
            'activity': activity_grade.grade if activity_grade and activity_grade.grade else '',
            'monthly': monthly_grade.grade if monthly_grade and monthly_grade.grade else '',
            'midterm': midterm_grade.grade if midterm_grade and midterm_grade.grade else '',
            'final': final_grade.grade if final_grade and final_grade.grade else '',
            'activity_notes': activity_grade.notes if activity_grade else '',
            'total': total
        })
    
    # إنشاء PDF باستخدام xhtml2pdf
    html_string = render_to_string('grade/print_grades.html', {
        'classroom': classroom,
        'subject': subject,
        'subject_display_name': subject_display_name,
        'students_data': students_data,
        'exam_types': exam_types
    })
    
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=grades_{classroom.name}_{subject.name}.pdf'
        return response
    
    return HttpResponse('Error generating PDF: %s' % pdf.err)
    
def select_subject(request, classroom_id):
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    subjects = classroom.classroomsubject_set.all()  
    
    # إضافة أسماء المعلمين للمواد
    subjects_with_teachers = []
    for subject in subjects:
        teacher_names = ", ".join([teacher.full_name for teacher in subject.subject.teachers.all()])
        if teacher_names:
            display_name = f"{subject.subject.name} ({teacher_names})"
        else:
            display_name = subject.subject.name
        subjects_with_teachers.append({
            'subject': subject,
            'display_name': display_name
        })
    
    return render(request, 'grade/select_subject.html', {
        'classroom': classroom,
        'subjects_with_teachers': subjects_with_teachers
    })
    
    


def custom_print_grades(request, classroom_id, subject_id):
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    
    # الحصول على أسماء المعلمين للمادة
    teacher_names = ", ".join([teacher.full_name for teacher in subject.teachers.all()])
    subject_display_name = f"{subject.name} ({teacher_names})" if teacher_names else subject.name
    
    grades = Grade.objects.filter(
        classroom=classroom,
        subject=subject
    ).select_related('student')
    
    # حساب المجموع لكل طالب
    students_data = []
    for student in classroom.students.all().order_by('full_name'):
        student_grades = grades.filter(student=student)
        activity_grade = student_grades.filter(exam_type='activity').first()
        monthly_grade = student_grades.filter(exam_type='monthly').first()
        midterm_grade = student_grades.filter(exam_type='midterm').first()
        final_grade = student_grades.filter(exam_type='final').first()
        
        # حساب المجموع
        total = 0
        if activity_grade and activity_grade.grade: 
            total += float(activity_grade.grade)
        if monthly_grade and monthly_grade.grade: 
            total += float(monthly_grade.grade)
        if midterm_grade and midterm_grade.grade: 
            total += float(midterm_grade.grade)
        if final_grade and final_grade.grade: 
            total += float(final_grade.grade)
            
        students_data.append({
            'student': student,
            'activity': activity_grade.grade if activity_grade and activity_grade.grade else '',
            'monthly': monthly_grade.grade if monthly_grade and monthly_grade.grade else '',
            'midterm': midterm_grade.grade if midterm_grade and midterm_grade.grade else '',
            'final': final_grade.grade if final_grade and final_grade.grade else '',
            'activity_notes': activity_grade.notes if activity_grade else '',
            'total': total
        })
    
    if request.method == 'POST':
        form = CustomPrintForm(request.POST)
        if form.is_valid():
            # إعداد البيانات للطباعة
            selected_tables = form.cleaned_data['tables']
            include_notes = form.cleaned_data['include_notes']
            include_signature = form.cleaned_data['include_signature']
            
            # إنشاء PDF
            html_string = render_to_string('grade/custom_print.html', {
                'classroom': classroom,
                'subject': subject,
                'subject_display_name': subject_display_name,
                'students_data': students_data,
                'selected_tables': selected_tables,
                'include_notes': include_notes,
                'include_signature': include_signature
            })
            
            result = io.BytesIO()
            pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
            
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename=grades_{classroom.name}_{subject.name}.pdf'
                return response
            
            return HttpResponse('Error generating PDF: %s' % pdf.err)
    else:
        form = CustomPrintForm()
    
    return render(request, 'grade/custom_print_options.html', {
        'classroom': classroom,
        'subject': subject,
        'subject_display_name': subject_display_name,
        'form': form
    }) 
    
    
def export_grades_excel(request, classroom_id, subject_id):
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    
    # الحصول على أسماء المعلمين للمادة
    teacher_names = ", ".join([teacher.full_name for teacher in subject.teachers.all()])
    subject_display_name = f"{subject.name} ({teacher_names})" if teacher_names else subject.name
    
    grades = Grade.objects.filter(
        classroom=classroom,
        subject=subject
    ).select_related('student')
    
    # حساب المجموع لكل طالب
    students_data = []
    for student in classroom.students.all().order_by('full_name'):
        student_grades = grades.filter(student=student)
        activity_grade = student_grades.filter(exam_type='activity').first()
        monthly_grade = student_grades.filter(exam_type='monthly').first()
        midterm_grade = student_grades.filter(exam_type='midterm').first()
        final_grade = student_grades.filter(exam_type='final').first()
        
        # حساب المجموع
        total = 0
        if activity_grade and activity_grade.grade: 
            total += float(activity_grade.grade)
        if monthly_grade and monthly_grade.grade: 
            total += float(monthly_grade.grade)
        if midterm_grade and midterm_grade.grade: 
            total += float(midterm_grade.grade)
        if final_grade and final_grade.grade: 
            total += float(final_grade.grade)
            
        students_data.append({
            'student': student,
            'activity': activity_grade.grade if activity_grade and activity_grade.grade else '',
            'monthly': monthly_grade.grade if monthly_grade and monthly_grade.grade else '',
            'midterm': midterm_grade.grade if midterm_grade and midterm_grade.grade else '',
            'final': final_grade.grade if final_grade and final_grade.grade else '',
            'activity_notes': activity_grade.notes if activity_grade else '',
            'total': total
        })
    
    # إنشاء ملف إكسل
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "العلامات"
    
    # كتابة العنوان
    ws['A1'] = f"علامات مادة {subject_display_name} - صف {classroom.name}"
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # كتابة العناوين
    headers = ['اسم الطالب', '1', '2', '3', 'نصفي', 'المجموع', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # كتابة البيانات
    for row, data in enumerate(students_data, 4):
        ws.cell(row=row, column=1, value=data['student'].full_name)
        ws.cell(row=row, column=2, value=data['activity'])
        ws.cell(row=row, column=3, value=data['monthly'])
        ws.cell(row=row, column=4, value=data['midterm'])
        ws.cell(row=row, column=5, value=data['final'])
        ws.cell(row=row, column=6, value=data['total'])
        ws.cell(row=row, column=7, value=data['activity_notes'])
    
    # ضبط عرض الأعمدة
    column_widths = [30, 10, 10, 10, 10, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # إعداد الاستجابة
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=grades_{classroom.name}_{subject.name}.xlsx'
    
    wb.save(response)
    return response    